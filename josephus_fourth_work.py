from openpyxl import load_workbook
import csv
import zipfile

excel_filepath = 'stu.list/stu.list.xlsx'
#filepath = 'stu.list/stu.list.csv'
#filepath = 'stu.list.zip'

class student:
    def __init__(self,name,id):
        assert(type(name) == str and type(id) == int)
        self.name = name
        self.id = id

    def __str__(self):
        return "{%s:%d}" % (self.name, self.id)
    __repr__ = __str__

class Reader:
    def __init__(self,filepath):
        self.filepath = filepath

class ExcelReader(Reader):
    def read(self):
        assert self.filepath.endswith(".xlsx")
        wb = load_workbook(self.filepath)
        sheet = wb.active
        stu_list = []
        for row in sheet.iter_rows():
            stu = student(name=row[0].value, id=row[1].value)
            stu_list.append(stu)
        return stu_list

class CsvReader(Reader):
    def read(self):
        assert self.filepath.endswith(".csv")
        with open(self.filepath) as f:
            f_csv = csv.reader(f)
            stu_list = []
            for row in f_csv:
                stu = student(name=row[0], id=row[1])
                stu_list.append(stu)
                return stu_list

class ZipReader(ExcelReader,CsvReader):
    def __init__(self,inter_filepath):
        self.inter_filepath = inter_filepath
    def read(self):
        assert self.filepath.endswith(".zip")
        with zipfile.ZipFile(self.filepath) as zfiles:
            namelist = zfiles.namelist()
            self.inter_filepath = namelist[0]
            zfiles.extractall()
            stu_list = []
            if self.inter_filepath.endswith('xlsx'):
                stu_list = self.read_excel()
            else: 
                stu_list = self.read_csv()
            return stu_list

class Josephus_Circle():
    def __init__(self,stu_list,step,start_id):
        assert len(stu_list) >= 0 
        assert step > 0
        assert 0 < start_id <= len(stu_list)
        self.step = step
        self.start_id = start_id
        self.result_list = []

    def __iter__(self):
        return self

    def __next__(self):
        if len(stu_list) == 1:
            return
        else:
            for i in range(len(stu_list)):
                if(self.start_id == stu_list[i].id):
                    index = i
                    break
            for i in range(len(stu_list)-1):
                index = (index + self.step) % len(stu_list) - 1
                self.result_list.append(stu_list[index])
                #print(stu_list[index].name,stu_list[index].id)
                del stu_list[index] 
                if index == -1: 
                    index = 0
            self.result_list.append(stu_list[0])
            #print(stu_list[0].name,stu_list[0].id)
            return self.result_list

if __name__ == '__main__':      
    #stu_list = ExcelReader(filepath).read()
    #stu_list = CsvReader(filepath).read()
    #stu_list = ZipReader(filepath).read()

    reader = ExcelReader(excel_filepath)
    for x in reader:
        print(x)
    josephus_circle = Josephus_Circle(stu_list, 3, 2)
    #print(josephus_circle)
    #for student in josephus_circle:
        #print(student)