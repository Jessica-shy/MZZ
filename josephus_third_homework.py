#1、从三种不同的文件来创建列表。eg：csv文件、excel文件、zip文件                                   
#2、用reader作为基类，可能有excel_reader等。用reader继承结构的方式，做约瑟夫环                                                           
#3、用for完成约瑟夫环容器的遍历

from openpyxl import load_workbook
import csv
import zipfile

excel_file_path = 'std.list/std.list.xlsx'
csv_file_path = 'std.list/std.list.csv'
zip_file_path = 'std.list.zip'

class student:
    def __init__(self,name,id):
        self.name = name
        self.id = id

class Reader:
    def __init__(self):
        pass

class ExcelReader(Reader):
    def __init__(self,file_path):
        self.path= file_path
    def read_excel(self):
        wb = load_workbook(self.path)
        sheet = wb.active
        std_list = []
        for row in sheet.iter_rows():
            std = student(name=row[0].value, id=row[1].value)
            std_list.append(std)
        return std_list

class CsvReader(Reader):
    def __init__(self,file_path):
        self.path= file_path
    def read_csv(self):
        with open(self.path) as f:
            f_csv = csv.reader(f)
            std_list = []
            for row in f_csv:
                std = student(name=row[0], id=row[1])
                std_list.append(std)
                return std_list

class ZipReader(ExcelReader):
    def __init__(self,file_path):
        self.path = file_path
    def read_zip(self):
        with zipfile.ZipFile(self.path) as zfiles:
            namelist = zfiles.namelist()
            self.path = namelist[0]
            zfiles.extractall()
            std_list = []
            std_list = self.read_excel()
            return std_list

def get_josephus(total_num,step,start_id):
    assert total_num > 0
    assert step > 0
    assert 0 < start_id <= total_num

    L = std_list.copy()
    if total_num == 1:
        return
    else:
        for i in range(len(std_list)):
            if(start_id == std_list[i].id):
                index = i
                break
        for i in range(total_num-1):
            index = (index + step) % len(L) - 1 #L[(上一次出局者的索引+k) % 上一次出局后的剩余人数 - 1]
            print(L[index].name,L[index].id)
            del L[index] #删除列表L中索引值为index的元素
            if index == -1: #只剩一个人(index < 0)
                index = 0
        print(L[0].name,L[0].id)

if __name__ == '__main__':      
    std_list = ExcelReader(excel_file_path).read_excel()
    #std_list = CsvReader(csv_file_path).read_csv()
    #std_list = ZipReader(zip_file_path).read_zip()
    
    get_josephus(10,2,3)
